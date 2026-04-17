#!/usr/bin/env python3
"""
1-Group: P&L & GL Financial Analysis Agent — Excel Builder

Builds a professional multi-sheet Excel workbook that cross-references P&L statements
with General Ledger detail to produce comprehensive expense breakdowns with vendor
attribution, COGS analysis, manpower analysis, and month-on-month variance tracking.

Usage:
    python build_pnl_gl_report.py --pnl pnl.xlsx --gl gl.xlsx [--output report.xlsx] [--industry hospitality]
    python build_pnl_gl_report.py --demo [--output demo_report.xlsx]
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from datetime import datetime, date
import argparse
import sys
import os
import re
from difflib import SequenceMatcher

# ═══════════════════════════════════════════════════════════════════════════════
# COLOUR PALETTE
# ═══════════════════════════════════════════════════════════════════════════════
C = {
    "hdr_dark":     "1B2A4A",
    "hdr_mid":      "2E4A7A",
    "hdr_light":    "4A7AB5",
    "accent":       "E8912D",
    "good":         "27AE60",
    "warn":         "F39C12",
    "bad":          "E74C3C",
    "bg_light":     "F4F6F9",
    "bg_white":     "FFFFFF",
    "bg_section":   "EBF0F8",
    "text_dark":    "1C1C1C",
    "text_blue":    "0000FF",   # Hardcoded inputs
    "text_black":   "000000",   # Formulas
    "text_green":   "008000",   # Cross-sheet refs
    "text_red":     "FF0000",   # External refs / flags
    "border":       "BDC3C7",
    "bg_yellow":    "FFFF00",   # Attention cells
    "bg_green_lt":  "E8F5E9",
    "bg_red_lt":    "FFEBEE",
    "bg_amber_lt":  "FFF8E1",
}

# ═══════════════════════════════════════════════════════════════════════════════
# STYLE HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
FONT_BASE = "Arial"

def hdr_font(color="FFFFFF", bold=True, size=10):
    return Font(name=FONT_BASE, bold=bold, color=color, size=size)

def data_font(color="000000", bold=False, size=10):
    return Font(name=FONT_BASE, bold=bold, color=color, size=size)

def cell_border(style="thin"):
    s = Side(style=style, color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def fill_color(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

FMT_CURRENCY = '$#,##0;($#,##0);"-"'
FMT_CURRENCY_2DP = '$#,##0.00;($#,##0.00);"-"'
FMT_PCT = '0.0%;(0.0%);"-"'
FMT_INT = '#,##0'
FMT_TEXT = '@'

def style_header_row(ws, row, max_col, bg=None, font_color="FFFFFF"):
    bg = bg or C["hdr_dark"]
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = hdr_font(color=font_color)
        cell.fill = fill_color(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border()

def style_data_cell(ws, row, col, fmt=None, bold=False, color="000000"):
    cell = ws.cell(row=row, column=col)
    cell.font = data_font(color=color, bold=bold)
    cell.border = cell_border()
    if fmt:
        cell.number_format = fmt
    return cell

def auto_width(ws, min_width=10, max_width=45):
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len

def write_section_header(ws, row, text, max_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = hdr_font(size=11)
    cell.fill = fill_color(C["hdr_mid"])
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).border = cell_border()

# ═══════════════════════════════════════════════════════════════════════════════
# VENDOR EXTRACTION ENGINE
# ═══════════════════════════════════════════════════════════════════════════════
TRANSACTION_PREFIXES = [
    r'^INV[-/\s]?\d+[-/\s]*\d*\s+', r'^PINV[-/\s]?\d+[-/\s]*\d*\s+',
    r'^SINV[-/\s]?\d+[-/\s]*\d*\s+',
    r'^AP[-/\s]?\d+\s+', r'^BILL[-/\s]?\d+[-/\s]*\d*\s+',
    r'^PV[-/\s]?\d+[-/\s]*\d*\s+',
    r'^PAY[-/\s]?\d+[-/\s]*\d*\s+', r'^PMT[-/\s]?\d+[-/\s]*\d*\s+',
    r'^JV[-/\s]?\d*\s*[-–]?\s*', r'^JE[-/\s]?\d+[-/\s]*\d*\s+',
    r'^MJE[-/\s]?\d*\s+', r'^ADJ[-/\s]?\d*\s+',
    r'^CN[-/\s]?\d+[-/\s]*\d*\s+', r'^DN[-/\s]?\d+[-/\s]*\d*\s+',
    r'^RCPT[-/\s]?\d+[-/\s]*\d*\s+',
    r'^GIRO\s*[-–]\s*', r'^TT[-/\s]?\d*\s+', r'^EFT[-/\s]?\d*\s+',
    r'^CHQ[-/\s]?\d*\s+', r'^DD\s+', r'^FAST\s*[-–]?\s*',
]

REFERENCE_PATTERNS = [
    r'\bREF\s*[:#]?\s*[A-Z0-9]+', r'\b(?:INV|PO|SO|DO|GRN)[-#]\d{3,}\b',
    r'#\d{3,}', r'\b\d{8,}\b', r'\b\d{2}[/-]\d{2}[/-]\d{2,4}\b',
    r'\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}\b',
    r'\bFY\d{2,4}\b', r'\bQ[1-4]\s+\d{4}\b',
]

PAYMENT_INDICATORS = [
    r'\b[Pp]ayment\s+(?:to|for|via)\s+', r'\b[Pp]aid\s+(?:to|for|via)\s+',
    r'\b(?:Bank\s+)?[Tt]ransfer\s+(?:to|for)\s+',
    r'\bGIRO\s+(?:Payment|Deduction)\s*', r'\bDirect\s+Debit\s*',
    r'\bBeing\s+(?:payment|reimbursement|accrual)\s+(?:for|of|to)\s+',
    r'\bAccrual\s+(?:for|of)\s+', r'\bReversal\s+(?:of|for)\s+',
    r'\bMonthly\s+(?:payment|charge|fee|subscription)\s*',
]

ENTITY_SUFFIXES = [
    r'\s+PTE\.?\s*LTD\.?$', r'\s+SDN\.?\s*BHD\.?$', r'\s+PTY\.?\s*LTD\.?$',
    r'\s+CO\.?\s*$', r'\s+CORP\.?$', r'\s+INC\.?$', r'\s+LLC\.?$',
    r'\s+LLP\.?$', r'\s+LTD\.?$', r'\s+LIMITED$', r'\s+PRIVATE\s+LIMITED$',
    r'\s+\(S\)$', r'\s+\(SG\)$', r'\s+SINGAPORE$',
]

SPECIAL_CATEGORIES = {
    'Petty Cash': [r'\bpetty\s*cash\b', r'\bPCF\b', r'\bcash\s+purchase\b'],
    'Sundry': [r'\bsundry\b', r'\bsundries\b', r'\bmiscellaneous\b', r'\bmisc\b'],
    'Internal/Payroll': [r'\bsalar(?:y|ies)\b', r'\bpayroll\b', r'\bwages?\b', r'\bCPF\b',
                         r'\bstaff\s+(?:meal|welfare|benefit)\b', r'\bbonus\b'],
    'Non-Cash': [r'\bdepreciation\b', r'\bamortisation\b', r'\bamortization\b',
                 r'\bprovision\b', r'\bwrite[\s-]?off\b', r'\bimpairment\b'],
    'Intercompany': [r'\binter[\s-]?company\b', r'\binter[\s-]?co\b', r'\b1[\s-]?group\b'],
    'Bank Charges': [r'\bbank\s+charge\b', r'\bservice\s+charge\b', r'\bmerchant\s+fee\b'],
}

def classify_special(description):
    desc_lower = description.lower()
    for category, patterns in SPECIAL_CATEGORIES.items():
        for pattern in patterns:
            if re.search(pattern, desc_lower):
                return category
    return None

def extract_vendor(description, contact_name=None):
    if contact_name and str(contact_name).strip() not in ('', 'nan', 'None', 'NaN'):
        return normalise_vendor(str(contact_name).strip())
    if not description or str(description).strip() in ('', 'nan', 'None', 'NaN'):
        return 'Unknown'
    text = str(description).strip()
    special = classify_special(text)
    if special:
        return special
    for p in TRANSACTION_PREFIXES:
        text = re.sub(p, '', text, flags=re.IGNORECASE)
    for p in REFERENCE_PATTERNS:
        text = re.sub(p, '', text, flags=re.IGNORECASE)
    for p in PAYMENT_INDICATORS:
        text = re.sub(p, '', text, flags=re.IGNORECASE)
    text = re.sub(r'\s*[-–]\s*(for\s+)?.*$', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\s*\$[\d,]+\.?\d*\s*$', '', text)
    text = re.sub(r'^[\s\-–—:;,\.]+|[\s\-–—:;,\.]+$', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return normalise_vendor(text) if len(text) > 1 else 'Unknown'

def normalise_vendor(name):
    n = name.upper().strip()
    for suffix in ENTITY_SUFFIXES:
        n = re.sub(suffix, '', n, flags=re.IGNORECASE)
    n = re.sub(r'[\s\.\,\-]+$', '', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n.title() if n else 'Unknown'

def fuzzy_group_vendors(vendor_list, threshold=0.85):
    groups = {}
    sorted_vendors = sorted(set(vendor_list), key=lambda x: (-len(x), x))
    for vendor in sorted_vendors:
        matched = False
        for canonical in groups:
            ratio = SequenceMatcher(None, vendor.upper(), canonical.upper()).ratio()
            if ratio >= threshold:
                groups[canonical].append(vendor)
                matched = True
                break
        if not matched:
            groups[vendor] = [vendor]
    return groups

# ═══════════════════════════════════════════════════════════════════════════════
# ACCOUNTING NUMBER PARSER
# ═══════════════════════════════════════════════════════════════════════════════
def parse_accounting_number(val):
    if pd.isna(val) or str(val).strip() in ('', '-', '—', '–'):
        return 0.0
    s = str(val).strip()
    negative = s.startswith('(') and s.endswith(')')
    s = s.replace('(', '').replace(')', '').replace('$', '').replace(',', '').replace(' ', '')
    try:
        result = float(s)
        return -result if negative else result
    except ValueError:
        return 0.0

# ═══════════════════════════════════════════════════════════════════════════════
# GL COLUMN DETECTOR
# ═══════════════════════════════════════════════════════════════════════════════
UNIVERSAL_COLUMN_MAP = {
    'account_code': ['Account Code', 'Account', 'Acct', 'Account No', 'Account Number',
                     'Account #', 'GL Account', 'Nominal Code', 'N/C', 'G/L Account',
                     'G/L Account No.'],
    'account_name': ['Account Name', 'Account Description', 'Nominal Name',
                     'G/L Account Name', 'Account: Name'],
    'date': ['Date', 'Transaction Date', 'Trans Date', 'Posting Date', 'Document Date',
             'Entry Date', 'Value Date'],
    'description': ['Description', 'Narration', 'Memo', 'Details', 'Narrative', 'Text',
                    'Line Description', 'Memo/Description', 'Particulars'],
    'vendor': ['Contact', 'Contact Name', 'Name', 'Vendor', 'Vendor Name', 'Supplier',
               'Supplier Name', 'Card Name', 'Payee', 'Entity', 'Business Partner'],
    'debit': ['Debit', 'Debit Amount', 'Dr', 'Dr Amount'],
    'credit': ['Credit', 'Credit Amount', 'Cr', 'Cr Amount'],
    'amount': ['Amount', 'Net Amount', 'Gross', 'Value', 'Balance'],
    'reference': ['Reference', 'Ref', 'Ref No.', 'Document Number', 'Doc. No.', 'Num'],
    'type': ['Type', 'Transaction Type', 'Source', 'Source Type', 'Document Type'],
}

def detect_gl_columns(df):
    mapping = {}
    headers = [str(c).strip() for c in df.columns]
    headers_upper = [h.upper() for h in headers]
    for role, candidates in UNIVERSAL_COLUMN_MAP.items():
        for candidate in candidates:
            if candidate.upper() in headers_upper:
                idx = headers_upper.index(candidate.upper())
                mapping[role] = headers[idx]
                break
    has_amounts = ('debit' in mapping and 'credit' in mapping) or 'amount' in mapping
    return mapping, has_amounts

# ═══════════════════════════════════════════════════════════════════════════════
# P&L CATEGORY CLASSIFIER
# ═══════════════════════════════════════════════════════════════════════════════
PNL_CATEGORIES = {
    'revenue': ['revenue', 'sales', 'income', 'turnover', 'fee income'],
    'cogs': ['cost of goods', 'cost of sales', 'cogs', 'direct cost', 'food cost',
             'beverage cost', 'bev cost', 'purchases', 'raw material'],
    'manpower': ['salary', 'salaries', 'wages', 'payroll', 'staff cost', 'manpower',
                 'cpf', 'pension', 'staff benefit', 'overtime', 'casual labour',
                 'training', 'recruitment', 'work permit', 'levy', 'bonus'],
    'occupancy': ['rent', 'lease', 'property tax', 'utilities', 'electricity', 'water', 'gas'],
    'admin': ['office', 'stationery', 'professional fee', 'legal', 'audit', 'insurance',
              'license', 'licence', 'permit', 'subscription'],
    'marketing': ['marketing', 'advertising', 'promotion', 'publicity', 'social media'],
    'operations': ['repair', 'maintenance', 'r&m', 'cleaning', 'hygiene', 'laundry',
                   'linen', 'pest control', 'security', 'transport', 'telecom',
                   'telephone', 'internet', 'it ', 'software', 'technology'],
    'depreciation': ['depreciation', 'amortisation', 'amortization'],
    'finance': ['interest', 'bank charge', 'merchant fee', 'finance charge', 'fx ',
                'foreign exchange', 'forex'],
}

def classify_pnl_line(name):
    name_lower = name.lower().strip()
    for cat, keywords in PNL_CATEGORIES.items():
        if any(kw in name_lower for kw in keywords):
            return cat
    return 'other_opex'

# ═══════════════════════════════════════════════════════════════════════════════
# INDUSTRY DEFAULTS
# ═══════════════════════════════════════════════════════════════════════════════
INDUSTRY_DEFAULTS = {
    'hospitality': {
        'food_cost_target': 0.30, 'bev_cost_target': 0.22, 'combined_cogs_target': 0.28,
        'manpower_target': 0.30, 'prime_cost_target': 0.60, 'rent_target': 0.12,
        'ebitda_target': 0.15, 'net_profit_target': 0.08,
        'var_pct_threshold': 0.15, 'var_abs_threshold': 5000,
    },
    'retail': {
        'food_cost_target': 0.0, 'bev_cost_target': 0.0, 'combined_cogs_target': 0.55,
        'manpower_target': 0.15, 'prime_cost_target': 0.70, 'rent_target': 0.10,
        'ebitda_target': 0.12, 'net_profit_target': 0.06,
        'var_pct_threshold': 0.15, 'var_abs_threshold': 5000,
    },
    'professional_services': {
        'food_cost_target': 0.0, 'bev_cost_target': 0.0, 'combined_cogs_target': 0.50,
        'manpower_target': 0.65, 'prime_cost_target': 1.0, 'rent_target': 0.05,
        'ebitda_target': 0.20, 'net_profit_target': 0.15,
        'var_pct_threshold': 0.15, 'var_abs_threshold': 10000,
    },
}

def get_defaults(industry='hospitality'):
    return INDUSTRY_DEFAULTS.get(industry, INDUSTRY_DEFAULTS['hospitality'])

# ═══════════════════════════════════════════════════════════════════════════════
# DEMO DATA GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════
def generate_demo_data():
    """Generate realistic demo P&L and GL data for a restaurant over 3 months."""
    months = ['Jan 2025', 'Feb 2025', 'Mar 2025']
    np.random.seed(42)

    pnl_data = {
        'Account Code': ['4000','4100','4200','','5000','5100','5200','','','6000','6100',
                         '6200','6300','6400','','','7000','7100','7200','7300','7400','7500',
                         '7600','7700','7800','7900','8000','8100','8200','','','9000','9100',
                         '9200','','',''],
        'Description': [
            'Food Revenue','Beverage Revenue - Alcoholic','Beverage Revenue - Non-Alcoholic',
            'TOTAL REVENUE',
            'Food Cost','Beverage Cost - Alcoholic','Beverage Cost - Non-Alcoholic',
            'TOTAL COGS','GROSS PROFIT',
            'Salaries & Wages','CPF Contributions','Staff Benefits','Overtime','Casual Labour',
            'TOTAL MANPOWER','',
            'Rent & Lease','Utilities','Repairs & Maintenance','Cleaning & Hygiene',
            'Marketing & Advertising','Insurance','Professional Fees','IT & Technology',
            'Depreciation','Bank Charges & Merchant Fees',
            'Laundry & Linen','Pest Control','Telecommunications',
            'TOTAL OPERATING EXPENSES','',
            'Interest Expense','Foreign Exchange Loss','Other Finance Charges',
            'TOTAL FINANCE COSTS','',
            'NET PROFIT'
        ],
    }

    base_rev = [320000, 95000, 45000]
    for i, month in enumerate(months):
        factor = 1 + np.random.uniform(-0.05, 0.08)
        rev = [int(v * factor * (1 + i * 0.02)) for v in base_rev]
        total_rev = sum(rev)
        food_cost = int(rev[0] * np.random.uniform(0.28, 0.33))
        bev_alc_cost = int(rev[1] * np.random.uniform(0.20, 0.25))
        bev_na_cost = int(rev[2] * np.random.uniform(0.12, 0.18))
        total_cogs = food_cost + bev_alc_cost + bev_na_cost
        gp = total_rev - total_cogs

        salary = int(total_rev * np.random.uniform(0.20, 0.24))
        cpf = int(salary * 0.17)
        benefits = int(total_rev * np.random.uniform(0.015, 0.025))
        ot = int(total_rev * np.random.uniform(0.01, 0.02))
        casual = int(total_rev * np.random.uniform(0.02, 0.04))
        total_manpower = salary + cpf + benefits + ot + casual

        rent = 38000
        utilities = int(np.random.uniform(12000, 16000))
        rm = int(np.random.uniform(4000, 8000))
        cleaning = int(np.random.uniform(3000, 6000))
        marketing = int(np.random.uniform(8000, 15000))
        insurance = 3500
        profees = int(np.random.uniform(2000, 5000))
        it_tech = int(np.random.uniform(3000, 5000))
        deprec = 12000
        bank = int(total_rev * np.random.uniform(0.015, 0.025))
        laundry = int(np.random.uniform(2000, 3500))
        pest = int(np.random.uniform(800, 1200))
        telecom = int(np.random.uniform(1500, 2500))
        total_opex = (rent + utilities + rm + cleaning + marketing + insurance + profees +
                      it_tech + deprec + bank + laundry + pest + telecom)

        interest = int(np.random.uniform(2000, 3500))
        fx = int(np.random.uniform(-500, 1000))
        other_fin = int(np.random.uniform(200, 800))
        total_fin = interest + fx + other_fin

        net_profit = gp - total_manpower - total_opex - total_fin

        pnl_data[month] = [
            rev[0], rev[1], rev[2], total_rev,
            food_cost, bev_alc_cost, bev_na_cost, total_cogs, gp,
            salary, cpf, benefits, ot, casual, total_manpower, '',
            rent, utilities, rm, cleaning, marketing, insurance, profees, it_tech,
            deprec, bank, laundry, pest, telecom, total_opex, '',
            interest, fx, other_fin, total_fin, '',
            net_profit
        ]

    df_pnl = pd.DataFrame(pnl_data)

    # --- Generate GL detail ---
    vendors = {
        '5000': [('Fresh Produce Sg', 0.35), ('Wang Kee Meats', 0.25),
                 ('Ocean Seafood Trading', 0.20), ('Daily Fresh Dairy', 0.12),
                 ('Petty Cash', 0.08)],
        '5100': [('Wine Connection Distribution', 0.40), ('Asia Pacific Breweries', 0.30),
                 ('Spirits Trading Co', 0.20), ('Craft Beer Sg', 0.10)],
        '5200': [('F&N Beverages', 0.45), ('Coca-Cola Sg', 0.30), ('Pokka Corp', 0.25)],
        '6000': [('Internal/Payroll', 1.0)],
        '6100': [('Internal/Payroll', 1.0)],
        '6200': [('Internal/Payroll', 1.0)],
        '6300': [('Internal/Payroll', 1.0)],
        '6400': [('Manpower Staffing Sg', 0.6), ('Temp Heroes Agency', 0.4)],
        '7000': [('Capitaland Mall Trust', 1.0)],
        '7100': [('Sp Services', 0.65), ('Pub Utilities Board', 0.25), ('City Gas', 0.10)],
        '7200': [('Aircon Experts Sg', 0.40), ('Plumbing Solutions', 0.30),
                 ('General Maintenance Co', 0.30)],
        '7300': [('CleanCo', 0.50), ('Hygiene Solutions Sg', 0.30), ('Petty Cash', 0.20)],
        '7400': [('Social Media Agency X', 0.45), ('Google Ads', 0.30),
                 ('Print Media Co', 0.25)],
        '7500': [('Aig Insurance Sg', 1.0)],
        '7600': [('Kpmg Sg', 0.50), ('Drew & Napier', 0.30), ('Hr Consultants Sg', 0.20)],
        '7700': [('Revel Systems', 0.40), ('Aws Singapore', 0.35), ('It Support Co', 0.25)],
        '7800': [('Non-Cash', 1.0)],
        '7900': [('Dbs Bank', 0.40), ('Stripe Sg', 0.35), ('Nets', 0.25)],
        '8000': [('Pressto Laundry', 0.60), ('Linen Supplies Sg', 0.40)],
        '8100': [('Rentokil Sg', 1.0)],
        '8200': [('Singtel', 0.60), ('Starhub', 0.40)],
        '9000': [('Dbs Bank', 0.70), ('Ocbc Bank', 0.30)],
        '9100': [('Fx Transaction', 1.0)],
        '9200': [('Bank Charges', 1.0)],
    }

    gl_rows = []
    for i, month in enumerate(months):
        month_num = i + 1
        pnl_values = dict(zip(pnl_data['Account Code'], pnl_data[month]))
        for acct_code, vendor_splits in vendors.items():
            total = pnl_values.get(acct_code, 0)
            if not total or total == '':
                continue
            total = abs(int(total))
            for vendor_name, pct in vendor_splits:
                vendor_amount = int(total * pct)
                n_txns = max(1, np.random.randint(2, 6))
                for t in range(n_txns):
                    txn_amt = vendor_amount // n_txns if t < n_txns - 1 else vendor_amount - (vendor_amount // n_txns) * (n_txns - 1)
                    day = np.random.randint(1, 28)
                    ref_num = np.random.randint(1000, 9999)
                    gl_rows.append({
                        'Account Code': acct_code,
                        'Account Name': dict(zip(pnl_data['Account Code'], pnl_data['Description'])).get(acct_code, ''),
                        'Date': f'2025-{month_num:02d}-{day:02d}',
                        'Description': f'INV-2025-{ref_num} {vendor_name}',
                        'Debit': txn_amt if txn_amt > 0 else 0,
                        'Credit': abs(txn_amt) if txn_amt < 0 else 0,
                        'Reference': f'INV-2025-{ref_num}',
                    })

    df_gl = pd.DataFrame(gl_rows)
    return df_pnl, df_gl, months

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ═══════════════════════════════════════════════════════════════════════════════

def build_assumptions_sheet(wb, defaults, entity_name, currency, periods):
    ws = wb.create_sheet("⚙️ Assumptions")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40

    ws.merge_cells("A1:C1")
    ws["A1"] = "P&L & GL ANALYSIS — ASSUMPTIONS & TARGETS"
    ws["A1"].font = hdr_font(size=13)
    ws["A1"].fill = fill_color(C["hdr_dark"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    row = 3
    assumptions = [
        ("Entity Name", entity_name, "Name of business entity"),
        ("Reporting Currency", currency, "Currency symbol used"),
        ("Periods Covered", ", ".join(periods), "Months/quarters in this analysis"),
        ("Industry", "Hospitality / F&B", "Used for benchmark selection"),
        ("", "", ""),
        ("— COGS TARGETS —", "", ""),
        ("Food Cost % Target", defaults['food_cost_target'], "Target food cost as % of food revenue"),
        ("Beverage Cost % Target", defaults['bev_cost_target'], "Target bev cost as % of bev revenue"),
        ("Combined COGS % Target", defaults['combined_cogs_target'], "Target combined COGS as % of total revenue"),
        ("", "", ""),
        ("— MANPOWER TARGETS —", "", ""),
        ("Manpower % of Revenue Target", defaults['manpower_target'], "Target total staff cost as % of revenue"),
        ("Prime Cost Target (COGS+Manpower)", defaults['prime_cost_target'], "COGS + Manpower should not exceed this % of revenue"),
        ("", "", ""),
        ("— PROFITABILITY TARGETS —", "", ""),
        ("Rent % Target", defaults['rent_target'], ""),
        ("EBITDA % Target", defaults['ebitda_target'], ""),
        ("Net Profit % Target", defaults['net_profit_target'], ""),
        ("", "", ""),
        ("— VARIANCE THRESHOLDS —", "", ""),
        ("MoM % Change Threshold", defaults['var_pct_threshold'], "Flag if MoM change exceeds this %"),
        ("MoM $ Change Threshold", defaults['var_abs_threshold'], "Flag if MoM absolute change exceeds this $"),
    ]

    for label, value, note in assumptions:
        if label.startswith("—"):
            write_section_header(ws, row, label, 3)
            row += 1
            continue
        ws.cell(row=row, column=1, value=label).font = data_font(bold=True)
        ws.cell(row=row, column=1).border = cell_border()
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.border = cell_border()
        if isinstance(value, float) and value < 1:
            val_cell.number_format = FMT_PCT
        elif isinstance(value, (int, float)) and value >= 1:
            val_cell.number_format = FMT_CURRENCY
        val_cell.font = data_font(color=C["text_blue"])  # Editable input
        ws.cell(row=row, column=3, value=note).font = data_font(color="888888", size=9)
        ws.cell(row=row, column=3).border = cell_border()
        row += 1

    return ws


def build_pnl_summary_sheet(wb, df_pnl, periods):
    ws = wb.create_sheet("📋 P&L Summary")
    ws.sheet_view.showGridLines = True

    headers = ['Account Code', 'Description'] + periods
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    style_header_row(ws, 1, len(headers))

    for row_idx, (_, row_data) in enumerate(df_pnl.iterrows(), 2):
        acct_code = row_data.get('Account Code', '')
        desc = row_data.get('Description', '')
        ws.cell(row=row_idx, column=1, value=acct_code).font = data_font()
        ws.cell(row=row_idx, column=1).border = cell_border()

        is_total = any(kw in str(desc).upper() for kw in ['TOTAL', 'GROSS PROFIT', 'NET PROFIT', 'EBITDA'])
        ws.cell(row=row_idx, column=2, value=desc).font = data_font(bold=is_total)
        ws.cell(row=row_idx, column=2).border = cell_border()
        if is_total:
            ws.cell(row=row_idx, column=2).fill = fill_color(C["bg_section"])

        for p_idx, period in enumerate(periods, 3):
            val = row_data.get(period, '')
            cell = ws.cell(row=row_idx, column=p_idx, value=val if val != '' else None)
            cell.number_format = FMT_CURRENCY
            cell.font = data_font(bold=is_total)
            cell.border = cell_border()
            if is_total:
                cell.fill = fill_color(C["bg_section"])

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40
    for i in range(3, 3 + len(periods)):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df_pnl) + 1}"
    return ws


def build_expense_breakdown_sheet(wb, expense_vendor_data, periods):
    ws = wb.create_sheet("🔍 Expense Breakdown")
    headers = ['P&L Line Item', 'Account Code', 'P&L Amount', 'GL Total',
               'Variance', 'Status', 'Vendor', 'Vendor Amount', '% of Line']
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    style_header_row(ws, 1, len(headers))

    row = 2
    for item in expense_vendor_data:
        pnl_line = item['pnl_line']
        acct_code = item['account_code']
        pnl_amt = item['pnl_amount']
        gl_total = item['gl_total']
        variance = pnl_amt - gl_total
        vendors = item['vendors']

        # First vendor row includes the P&L line info
        first = True
        for vendor_name, vendor_amt in vendors:
            pct = vendor_amt / pnl_amt if pnl_amt != 0 else 0
            ws.cell(row=row, column=1, value=pnl_line if first else '').font = data_font(bold=first)
            ws.cell(row=row, column=2, value=acct_code if first else '').font = data_font()
            c_pnl = ws.cell(row=row, column=3, value=pnl_amt if first else None)
            c_pnl.number_format = FMT_CURRENCY
            c_gl = ws.cell(row=row, column=4, value=gl_total if first else None)
            c_gl.number_format = FMT_CURRENCY
            c_var = ws.cell(row=row, column=5, value=variance if first else None)
            c_var.number_format = FMT_CURRENCY
            if first:
                status = '✓ Ties' if abs(variance) < 1 else f'⚠ Variance: ${abs(variance):,.0f}'
                c_status = ws.cell(row=row, column=6, value=status)
                c_status.font = data_font(color=C["good"] if abs(variance) < 1 else C["bad"])
            ws.cell(row=row, column=7, value=vendor_name).font = data_font()
            ws.cell(row=row, column=8, value=vendor_amt).number_format = FMT_CURRENCY
            ws.cell(row=row, column=9, value=pct).number_format = FMT_PCT

            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).border = cell_border()
            first = False
            row += 1
        row += 1  # Blank row between expense lines

    auto_width(ws)
    ws.freeze_panes = "A2"
    return ws


def build_cogs_sheet(wb, cogs_data, periods, defaults):
    ws = wb.create_sheet("🥩 COGS Analysis")
    ws.merge_cells("A1:H1")
    ws["A1"] = "COST OF GOODS SOLD — ANALYSIS BY CATEGORY & VENDOR"
    ws["A1"].font = hdr_font(size=12)
    ws["A1"].fill = fill_color(C["hdr_dark"])
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ['Category', 'Vendor', 'Total Spend', '% of Category', '% of Revenue',
               'Target %', 'Variance vs Target', 'Status']
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=3, column=col_idx, value=h)
    style_header_row(ws, 3, len(headers))

    row = 4
    for category in cogs_data:
        write_section_header(ws, row, category['name'], len(headers))
        row += 1
        for vendor_name, amount in category['vendors']:
            pct_cat = amount / category['total'] if category['total'] else 0
            ws.cell(row=row, column=1, value='')
            ws.cell(row=row, column=2, value=vendor_name).font = data_font()
            ws.cell(row=row, column=3, value=amount).number_format = FMT_CURRENCY
            ws.cell(row=row, column=4, value=pct_cat).number_format = FMT_PCT
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).border = cell_border()
            row += 1
        # Category total row
        ws.cell(row=row, column=1, value=f"TOTAL {category['name'].upper()}").font = data_font(bold=True)
        ws.cell(row=row, column=3, value=category['total']).number_format = FMT_CURRENCY
        ws.cell(row=row, column=3).font = data_font(bold=True)
        if category.get('pct_revenue'):
            ws.cell(row=row, column=5, value=category['pct_revenue']).number_format = FMT_PCT
        if category.get('target'):
            ws.cell(row=row, column=6, value=category['target']).number_format = FMT_PCT
            ws.cell(row=row, column=6).font = data_font(color=C["text_blue"])
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = cell_border()
            ws.cell(row=row, column=c).fill = fill_color(C["bg_section"])
        row += 2

    auto_width(ws)
    return ws


def build_manpower_sheet(wb, manpower_data, total_revenue, defaults):
    ws = wb.create_sheet("👥 Manpower Analysis")
    ws.merge_cells("A1:F1")
    ws["A1"] = "MANPOWER COST ANALYSIS"
    ws["A1"].font = hdr_font(size=12)
    ws["A1"].fill = fill_color(C["hdr_dark"])
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ['Cost Component', 'Total Amount', '% of Total Manpower', '% of Revenue', 'Vendors/Agencies', 'Notes']
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=3, column=col_idx, value=h)
    style_header_row(ws, 3, len(headers))

    row = 4
    total_manpower = sum(item['amount'] for item in manpower_data)
    for item in manpower_data:
        ws.cell(row=row, column=1, value=item['component']).font = data_font()
        ws.cell(row=row, column=2, value=item['amount']).number_format = FMT_CURRENCY
        pct_mp = item['amount'] / total_manpower if total_manpower else 0
        pct_rev = item['amount'] / total_revenue if total_revenue else 0
        ws.cell(row=row, column=3, value=pct_mp).number_format = FMT_PCT
        ws.cell(row=row, column=4, value=pct_rev).number_format = FMT_PCT
        ws.cell(row=row, column=5, value=item.get('vendors', 'Internal')).font = data_font()
        ws.cell(row=row, column=6, value=item.get('notes', '')).font = data_font(color="888888", size=9)
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = cell_border()
        row += 1

    # Total row
    ws.cell(row=row, column=1, value='TOTAL MANPOWER').font = data_font(bold=True)
    ws.cell(row=row, column=2, value=total_manpower).number_format = FMT_CURRENCY
    ws.cell(row=row, column=2).font = data_font(bold=True)
    ws.cell(row=row, column=4, value=total_manpower / total_revenue if total_revenue else 0).number_format = FMT_PCT
    ws.cell(row=row, column=4).font = data_font(bold=True)
    for c in range(1, len(headers) + 1):
        ws.cell(row=row, column=c).border = cell_border()
        ws.cell(row=row, column=c).fill = fill_color(C["bg_section"])

    row += 2
    ws.cell(row=row, column=1, value='Manpower Target (% of Revenue)').font = data_font(bold=True)
    ws.cell(row=row, column=2, value=defaults['manpower_target']).number_format = FMT_PCT
    ws.cell(row=row, column=2).font = data_font(color=C["text_blue"])

    auto_width(ws)
    return ws


def build_mom_variance_sheet(wb, variance_data, periods):
    ws = wb.create_sheet("📈 MoM Variance")
    ws.merge_cells("A1:K1")
    ws["A1"] = "MONTH-ON-MONTH VARIANCE ANALYSIS"
    ws["A1"].font = hdr_font(size=12)
    ws["A1"].fill = fill_color(C["hdr_dark"])
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ['Expense Line'] + periods + ['MoM $ Change', 'MoM % Change', 'Avg', 'Flag', 'Variance Commentary']
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=3, column=col_idx, value=h)
    style_header_row(ws, 3, len(headers))

    row = 4
    for item in variance_data:
        ws.cell(row=row, column=1, value=item['line']).font = data_font(bold=True)
        ws.cell(row=row, column=1).border = cell_border()
        for p_idx, val in enumerate(item['values'], 2):
            c = ws.cell(row=row, column=p_idx, value=val)
            c.number_format = FMT_CURRENCY
            c.border = cell_border()
        col_offset = 2 + len(periods)
        ws.cell(row=row, column=col_offset, value=item.get('mom_dollar', 0)).number_format = FMT_CURRENCY
        ws.cell(row=row, column=col_offset).border = cell_border()
        ws.cell(row=row, column=col_offset + 1, value=item.get('mom_pct', 0)).number_format = FMT_PCT
        ws.cell(row=row, column=col_offset + 1).border = cell_border()
        ws.cell(row=row, column=col_offset + 2, value=item.get('avg', 0)).number_format = FMT_CURRENCY
        ws.cell(row=row, column=col_offset + 2).border = cell_border()
        flag = item.get('flag', '')
        flag_cell = ws.cell(row=row, column=col_offset + 3, value=flag)
        flag_cell.font = data_font(color=C["bad"] if flag else C["good"])
        flag_cell.border = cell_border()
        ws.cell(row=row, column=col_offset + 4, value=item.get('commentary', '')).font = data_font(size=9)
        ws.cell(row=row, column=col_offset + 4).border = cell_border()

        if flag:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = fill_color(C["bg_red_lt"])
        row += 1

    auto_width(ws)
    ws.freeze_panes = "B4"
    return ws


def build_vendor_summary_sheet(wb, all_vendors):
    ws = wb.create_sheet("🏭 Vendor Summary")
    headers = ['Rank', 'Vendor Name', 'Total Spend', '# Transactions', 'Expense Categories',
               'Avg Transaction', 'Cumulative %']
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    style_header_row(ws, 1, len(headers))

    sorted_vendors = sorted(all_vendors.items(), key=lambda x: -x[1]['total'])
    grand_total = sum(v['total'] for _, v in sorted_vendors) or 1
    cumulative = 0

    for rank, (vendor, data) in enumerate(sorted_vendors, 1):
        row = rank + 1
        cumulative += data['total']
        ws.cell(row=row, column=1, value=rank).font = data_font()
        ws.cell(row=row, column=2, value=vendor).font = data_font(bold=(rank <= 10))
        ws.cell(row=row, column=3, value=data['total']).number_format = FMT_CURRENCY
        ws.cell(row=row, column=4, value=data['count']).number_format = FMT_INT
        ws.cell(row=row, column=5, value=', '.join(data.get('categories', []))).font = data_font(size=9)
        avg_txn = data['total'] / data['count'] if data['count'] else 0
        ws.cell(row=row, column=6, value=avg_txn).number_format = FMT_CURRENCY
        ws.cell(row=row, column=7, value=cumulative / grand_total).number_format = FMT_PCT
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = cell_border()
        if cumulative / grand_total <= 0.80:
            ws.cell(row=row, column=2).fill = fill_color(C["bg_amber_lt"])

    auto_width(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(sorted_vendors) + 1}"
    return ws


def build_gl_detail_sheet(wb, df_gl):
    ws = wb.create_sheet("📒 GL Detail")
    headers = list(df_gl.columns) + ['Vendor (Extracted)']
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    style_header_row(ws, 1, len(headers))

    for row_idx, (_, gl_row) in enumerate(df_gl.iterrows(), 2):
        for col_idx, col_name in enumerate(df_gl.columns, 1):
            val = gl_row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = cell_border()
            if col_name in ('Debit', 'Credit'):
                cell.number_format = FMT_CURRENCY
        # Add extracted vendor column
        vendor = extract_vendor(str(gl_row.get('Description', '')))
        ws.cell(row=row_idx, column=len(df_gl.columns) + 1, value=vendor).border = cell_border()

    auto_width(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df_gl) + 1}"
    return ws


def build_notes_sheet(wb, notes):
    ws = wb.create_sheet("📝 Notes & Methodology")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 80

    ws["A1"] = "NOTES & METHODOLOGY"
    ws["A1"].font = hdr_font(size=13)
    ws["A1"].fill = fill_color(C["hdr_dark"])
    ws["A1"].alignment = Alignment(horizontal="center")

    row = 3
    for section, items in notes.items():
        ws.cell(row=row, column=1, value=section).font = data_font(bold=True, color=C["hdr_mid"])
        row += 1
        for item in items:
            ws.cell(row=row, column=1, value=f"  • {item}").font = data_font(size=9)
            row += 1
        row += 1

    return ws


def build_dashboard_sheet(wb, summary_kpis, top_vendors, periods):
    ws = wb.create_sheet("📊 Dashboard")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "P&L & GL ANALYSIS — EXECUTIVE DASHBOARD"
    ws["A1"].font = hdr_font(size=14)
    ws["A1"].fill = fill_color(C["hdr_dark"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # KPI Cards
    row = 3
    ws.cell(row=row, column=1, value="KEY PERFORMANCE INDICATORS").font = hdr_font(color=C["hdr_dark"], size=11)
    row = 4
    kpi_headers = ['Metric', 'Value', '% of Revenue', 'Target', 'Status']
    for col_idx, h in enumerate(kpi_headers, 1):
        ws.cell(row=row, column=col_idx, value=h)
    style_header_row(ws, row, len(kpi_headers))

    row = 5
    for kpi in summary_kpis:
        ws.cell(row=row, column=1, value=kpi['metric']).font = data_font(bold=True)
        ws.cell(row=row, column=2, value=kpi['value']).number_format = FMT_CURRENCY
        if kpi.get('pct_revenue') is not None:
            ws.cell(row=row, column=3, value=kpi['pct_revenue']).number_format = FMT_PCT
        if kpi.get('target') is not None:
            ws.cell(row=row, column=4, value=kpi['target']).number_format = FMT_PCT
            ws.cell(row=row, column=4).font = data_font(color=C["text_blue"])
        status = kpi.get('status', '')
        status_cell = ws.cell(row=row, column=5, value=status)
        color_map = {'🟢': C["good"], '🟡': C["warn"], '🔴': C["bad"]}
        for emoji, clr in color_map.items():
            if emoji in status:
                status_cell.font = data_font(color=clr, bold=True)
                break
        for c in range(1, len(kpi_headers) + 1):
            ws.cell(row=row, column=c).border = cell_border()
        row += 1

    # Top Vendors
    row += 2
    ws.cell(row=row, column=1, value="TOP 10 VENDORS BY SPEND").font = hdr_font(color=C["hdr_dark"], size=11)
    row += 1
    vendor_headers = ['Rank', 'Vendor', 'Total Spend', '% of Total Expenses']
    for col_idx, h in enumerate(vendor_headers, 1):
        ws.cell(row=row, column=col_idx, value=h)
    style_header_row(ws, row, len(vendor_headers))
    row += 1

    for rank, (name, spend, pct) in enumerate(top_vendors[:10], 1):
        ws.cell(row=row, column=1, value=rank).font = data_font()
        ws.cell(row=row, column=2, value=name).font = data_font(bold=True)
        ws.cell(row=row, column=3, value=spend).number_format = FMT_CURRENCY
        ws.cell(row=row, column=4, value=pct).number_format = FMT_PCT
        for c in range(1, len(vendor_headers) + 1):
            ws.cell(row=row, column=c).border = cell_border()
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16

    # Move Dashboard to first position
    wb.move_sheet(ws, offset=-len(wb.sheetnames) + 1)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

def build_report(df_pnl, df_gl, periods, entity_name="1-Group", currency="$",
                 industry="hospitality", output_path="pnl_gl_analysis.xlsx"):
    """Build the complete P&L & GL Analysis workbook."""
    defaults = get_defaults(industry)
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # --- Process GL data ---
    df_gl['Vendor_Extracted'] = df_gl.apply(
        lambda r: extract_vendor(str(r.get('Description', '')),
                                 r.get('Contact', r.get('Name', r.get('Vendor', None)))),
        axis=1
    )
    if 'Debit' in df_gl.columns and 'Credit' in df_gl.columns:
        df_gl['Net'] = df_gl['Debit'].apply(parse_accounting_number) - df_gl['Credit'].apply(parse_accounting_number)
    elif 'Amount' in df_gl.columns:
        df_gl['Net'] = df_gl['Amount'].apply(parse_accounting_number)

    # --- Build account → vendor mapping ---
    gl_by_account = df_gl.groupby('Account Code').agg(
        gl_total=('Net', 'sum'),
        txn_count=('Net', 'count')
    ).reset_index()

    vendor_by_account = df_gl.groupby(['Account Code', 'Vendor_Extracted']).agg(
        vendor_total=('Net', 'sum'),
        vendor_count=('Net', 'count')
    ).reset_index()

    # --- Build expense breakdown data ---
    expense_vendor_data = []
    all_vendors = {}

    pnl_lines = df_pnl[df_pnl['Account Code'].notna() & (df_pnl['Account Code'] != '')]
    for _, pnl_row in pnl_lines.iterrows():
        acct_code = str(pnl_row['Account Code']).strip()
        desc = str(pnl_row['Description']).strip()
        # Sum P&L amounts across all periods
        pnl_total = sum(parse_accounting_number(pnl_row.get(p, 0)) for p in periods)

        if pnl_total == 0 or any(kw in desc.upper() for kw in ['TOTAL', 'GROSS PROFIT', 'NET PROFIT', 'EBITDA']):
            continue

        # Get GL vendors for this account
        acct_vendors = vendor_by_account[vendor_by_account['Account Code'] == acct_code]
        gl_total = gl_by_account[gl_by_account['Account Code'] == acct_code]['gl_total'].sum()

        vendors = []
        for _, vrow in acct_vendors.iterrows():
            v_name = vrow['Vendor_Extracted']
            v_amt = abs(vrow['vendor_total'])
            vendors.append((v_name, v_amt))

            # Track for vendor summary
            if v_name not in all_vendors:
                all_vendors[v_name] = {'total': 0, 'count': 0, 'categories': set()}
            all_vendors[v_name]['total'] += v_amt
            all_vendors[v_name]['count'] += vrow['vendor_count']
            all_vendors[v_name]['categories'].add(classify_pnl_line(desc))

        vendors.sort(key=lambda x: -x[1])
        expense_vendor_data.append({
            'pnl_line': desc,
            'account_code': acct_code,
            'pnl_amount': abs(pnl_total),
            'gl_total': abs(gl_total),
            'vendors': vendors,
            'category': classify_pnl_line(desc),
        })

    # Convert vendor categories from sets to lists
    for v in all_vendors.values():
        v['categories'] = sorted(v['categories'])

    # --- Calculate summary KPIs ---
    pnl_dict = {}
    for _, row in df_pnl.iterrows():
        desc = str(row.get('Description', '')).strip()
        total = sum(parse_accounting_number(row.get(p, 0)) for p in periods)
        pnl_dict[desc] = total

    total_revenue = pnl_dict.get('TOTAL REVENUE', 0)
    total_cogs = abs(pnl_dict.get('TOTAL COGS', 0))
    gross_profit = pnl_dict.get('GROSS PROFIT', 0)
    total_manpower = abs(pnl_dict.get('TOTAL MANPOWER', 0))
    total_opex = abs(pnl_dict.get('TOTAL OPERATING EXPENSES', 0))
    total_finance = abs(pnl_dict.get('TOTAL FINANCE COSTS', 0))
    net_profit = pnl_dict.get('NET PROFIT', 0)

    def traffic_light(actual, target, invert=False):
        if invert:
            if actual >= target: return '🟢 On Target'
            elif actual >= target - 0.03: return '🟡 Watch'
            else: return '🔴 Below Target'
        else:
            if actual <= target: return '🟢 On Target'
            elif actual <= target + 0.03: return '🟡 Watch'
            else: return '🔴 Over Target'

    cogs_pct = total_cogs / total_revenue if total_revenue else 0
    manpower_pct = total_manpower / total_revenue if total_revenue else 0
    gp_pct = gross_profit / total_revenue if total_revenue else 0
    np_pct = net_profit / total_revenue if total_revenue else 0

    summary_kpis = [
        {'metric': 'Total Revenue', 'value': total_revenue, 'pct_revenue': None, 'target': None, 'status': ''},
        {'metric': 'Total COGS', 'value': total_cogs, 'pct_revenue': cogs_pct,
         'target': defaults['combined_cogs_target'], 'status': traffic_light(cogs_pct, defaults['combined_cogs_target'])},
        {'metric': 'Gross Profit', 'value': gross_profit, 'pct_revenue': gp_pct, 'target': None,
         'status': traffic_light(gp_pct, 0.65, invert=True)},
        {'metric': 'Total Manpower', 'value': total_manpower, 'pct_revenue': manpower_pct,
         'target': defaults['manpower_target'], 'status': traffic_light(manpower_pct, defaults['manpower_target'])},
        {'metric': 'Total Operating Expenses', 'value': total_opex, 'pct_revenue': total_opex / total_revenue if total_revenue else 0,
         'target': None, 'status': ''},
        {'metric': 'Total Finance Costs', 'value': total_finance, 'pct_revenue': total_finance / total_revenue if total_revenue else 0,
         'target': None, 'status': ''},
        {'metric': 'Net Profit', 'value': net_profit, 'pct_revenue': np_pct,
         'target': defaults['net_profit_target'], 'status': traffic_light(np_pct, defaults['net_profit_target'], invert=True)},
    ]

    # Top vendors for dashboard
    sorted_v = sorted(all_vendors.items(), key=lambda x: -x[1]['total'])
    total_expenses = sum(v['total'] for _, v in sorted_v) or 1
    top_vendors = [(name, data['total'], data['total'] / total_expenses) for name, data in sorted_v[:10]]

    # --- Build COGS breakdown ---
    cogs_items = [e for e in expense_vendor_data if e['category'] == 'cogs']
    cogs_data = []
    food_items = [e for e in cogs_items if 'food' in e['pnl_line'].lower()]
    bev_alc = [e for e in cogs_items if 'alcoholic' in e['pnl_line'].lower() or
               ('beverage' in e['pnl_line'].lower() and 'non' not in e['pnl_line'].lower())]
    bev_na = [e for e in cogs_items if 'non-alcoholic' in e['pnl_line'].lower() or 'non alcoholic' in e['pnl_line'].lower()]
    other_cogs = [e for e in cogs_items if e not in food_items + bev_alc + bev_na]

    for label, items, target_key in [
        ('Food Cost', food_items, 'food_cost_target'),
        ('Beverage Cost — Alcoholic', bev_alc, 'bev_cost_target'),
        ('Beverage Cost — Non-Alcoholic', bev_na, None),
        ('Other Direct Costs', other_cogs, None),
    ]:
        all_v = []
        total = 0
        for item in items:
            all_v.extend(item['vendors'])
            total += item['pnl_amount']
        cogs_data.append({
            'name': label,
            'vendors': sorted(all_v, key=lambda x: -x[1]),
            'total': total,
            'pct_revenue': total / total_revenue if total_revenue else 0,
            'target': defaults.get(target_key) if target_key else None,
        })

    # --- Build manpower breakdown ---
    manpower_items = [e for e in expense_vendor_data if e['category'] == 'manpower']
    manpower_data = []
    for item in manpower_items:
        vendor_str = ', '.join(v[0] for v in item['vendors'][:3])
        manpower_data.append({
            'component': item['pnl_line'],
            'amount': item['pnl_amount'],
            'vendors': vendor_str if vendor_str else 'Internal',
            'notes': '',
        })

    # --- Build MoM variance ---
    variance_data = []
    for item in expense_vendor_data:
        values = []
        for p in periods:
            pnl_row = df_pnl[df_pnl['Account Code'] == item['account_code']]
            if not pnl_row.empty:
                values.append(abs(parse_accounting_number(pnl_row.iloc[0].get(p, 0))))
            else:
                values.append(0)

        if len(values) >= 2:
            mom_dollar = values[-1] - values[-2]
            mom_pct = mom_dollar / values[-2] if values[-2] != 0 else 0
            avg_val = np.mean(values)
            flag = ''
            commentary = ''
            if abs(mom_pct) > defaults['var_pct_threshold'] or abs(mom_dollar) > defaults['var_abs_threshold']:
                flag = '⚠ FLAG'
                direction = '↑ increased' if mom_dollar > 0 else '↓ decreased'
                commentary = f"{item['pnl_line']} {direction} {abs(mom_pct):.0%} MoM (${abs(mom_dollar):,.0f})"
                if item['vendors']:
                    top_v = item['vendors'][0][0]
                    commentary += f" — largest vendor: {top_v}"
        else:
            mom_dollar, mom_pct, avg_val, flag, commentary = 0, 0, values[0] if values else 0, '', ''

        variance_data.append({
            'line': item['pnl_line'],
            'values': values,
            'mom_dollar': mom_dollar,
            'mom_pct': mom_pct,
            'avg': avg_val,
            'flag': flag,
            'commentary': commentary,
        })

    # --- Build notes ---
    notes = {
        'Mapping Methodology': [
            'GL accounts were matched to P&L lines using Account Code direct join',
            'Vendor names extracted from GL Description field using pattern matching',
            'Vendor name variants consolidated using fuzzy matching (85% threshold)',
        ],
        'Data Sources': [
            f'P&L report covering periods: {", ".join(periods)}',
            f'GL detail covering periods: {", ".join(periods)}',
            f'Entity: {entity_name}',
        ],
        'Caveats': [
            'Special categories (Petty Cash, Internal/Payroll, Non-Cash) are flagged but not attributed to external vendors',
            'Vendor name consolidation is approximate — review the Vendor Summary sheet for accuracy',
            'MoM variance commentary is auto-generated and may require manual review',
        ],
    }

    # Unmatched accounts
    all_pnl_codes = set(df_pnl['Account Code'].dropna().astype(str).str.strip())
    all_gl_codes = set(df_gl['Account Code'].dropna().astype(str).str.strip())
    gl_only = all_gl_codes - all_pnl_codes - {''}
    pnl_only = all_pnl_codes - all_gl_codes - {''}
    if gl_only:
        notes['GL Accounts Not in P&L'] = [f'Account {code}' for code in sorted(gl_only)]
    if pnl_only:
        notes['P&L Lines Without GL Detail'] = [f'Account {code}' for code in sorted(pnl_only)]

    # ═══ BUILD ALL SHEETS ═══
    build_assumptions_sheet(wb, defaults, entity_name, currency, periods)
    build_pnl_summary_sheet(wb, df_pnl, periods)
    build_expense_breakdown_sheet(wb, expense_vendor_data, periods)
    build_cogs_sheet(wb, cogs_data, periods, defaults)
    build_manpower_sheet(wb, manpower_data, total_revenue, defaults)

    # OpEx and Finance sheets follow same pattern as COGS
    opex_items = [e for e in expense_vendor_data if e['category'] in
                  ('occupancy', 'admin', 'marketing', 'operations', 'depreciation', 'other_opex')]
    opex_data = []
    for item in opex_items:
        opex_data.append({
            'name': item['pnl_line'],
            'vendors': item['vendors'],
            'total': item['pnl_amount'],
            'pct_revenue': item['pnl_amount'] / total_revenue if total_revenue else 0,
            'target': None,
        })
    ws_opex = wb.create_sheet("🏢 OpEx Analysis")
    ws_opex.merge_cells("A1:H1")
    ws_opex["A1"] = "OPERATING OVERHEAD EXPENSES — VENDOR BREAKDOWN"
    ws_opex["A1"].font = hdr_font(size=12)
    ws_opex["A1"].fill = fill_color(C["hdr_dark"])
    opex_headers = ['Expense Line', 'Vendor', 'Amount', '% of Line', '% of Revenue']
    for ci, h in enumerate(opex_headers, 1):
        ws_opex.cell(row=3, column=ci, value=h)
    style_header_row(ws_opex, 3, len(opex_headers))
    orow = 4
    for item in opex_data:
        for vi, (vn, va) in enumerate(item['vendors']):
            ws_opex.cell(row=orow, column=1, value=item['name'] if vi == 0 else '').font = data_font(bold=(vi==0))
            ws_opex.cell(row=orow, column=2, value=vn).font = data_font()
            ws_opex.cell(row=orow, column=3, value=va).number_format = FMT_CURRENCY
            pct_line = va / item['total'] if item['total'] else 0
            ws_opex.cell(row=orow, column=4, value=pct_line).number_format = FMT_PCT
            pct_rev = va / total_revenue if total_revenue else 0
            ws_opex.cell(row=orow, column=5, value=pct_rev).number_format = FMT_PCT
            for c in range(1, len(opex_headers) + 1):
                ws_opex.cell(row=orow, column=c).border = cell_border()
            orow += 1
        orow += 1
    auto_width(ws_opex)

    # Finance Costs sheet
    fin_items = [e for e in expense_vendor_data if e['category'] == 'finance']
    ws_fin = wb.create_sheet("💰 Finance Costs")
    ws_fin.merge_cells("A1:F1")
    ws_fin["A1"] = "FINANCE & NON-OPERATING COSTS"
    ws_fin["A1"].font = hdr_font(size=12)
    ws_fin["A1"].fill = fill_color(C["hdr_dark"])
    fin_headers = ['Cost Item', 'Counterparty', 'Amount', '% of Total Finance']
    for ci, h in enumerate(fin_headers, 1):
        ws_fin.cell(row=3, column=ci, value=h)
    style_header_row(ws_fin, 3, len(fin_headers))
    frow = 4
    total_fin_costs = sum(item['pnl_amount'] for item in fin_items) or 1
    for item in fin_items:
        for vi, (vn, va) in enumerate(item['vendors']):
            ws_fin.cell(row=frow, column=1, value=item['pnl_line'] if vi == 0 else '').font = data_font(bold=(vi==0))
            ws_fin.cell(row=frow, column=2, value=vn).font = data_font()
            ws_fin.cell(row=frow, column=3, value=va).number_format = FMT_CURRENCY
            ws_fin.cell(row=frow, column=4, value=va / total_fin_costs).number_format = FMT_PCT
            for c in range(1, len(fin_headers) + 1):
                ws_fin.cell(row=frow, column=c).border = cell_border()
            frow += 1
    auto_width(ws_fin)

    build_mom_variance_sheet(wb, variance_data, periods)
    build_vendor_summary_sheet(wb, all_vendors)
    build_gl_detail_sheet(wb, df_gl)
    build_notes_sheet(wb, notes)
    build_dashboard_sheet(wb, summary_kpis, top_vendors, periods)

    wb.save(output_path)
    return output_path, summary_kpis, top_vendors, variance_data


# ═══════════════════════════════════════════════════════════════════════════════
# CLI ENTRYPOINT
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description='P&L & GL Financial Analysis Agent')
    parser.add_argument('--pnl', help='Path to P&L file (xlsx/csv)')
    parser.add_argument('--gl', help='Path to GL detail file (xlsx/csv)')
    parser.add_argument('--output', default='pnl_gl_analysis.xlsx', help='Output xlsx path')
    parser.add_argument('--entity', default='1-Group', help='Entity name')
    parser.add_argument('--currency', default='$', help='Currency symbol')
    parser.add_argument('--industry', default='hospitality', choices=['hospitality', 'retail', 'professional_services'])
    parser.add_argument('--demo', action='store_true', help='Generate demo data')
    args = parser.parse_args()

    if args.demo:
        print("Generating demo data...")
        df_pnl, df_gl, periods = generate_demo_data()
    elif args.pnl and args.gl:
        print(f"Loading P&L: {args.pnl}")
        df_pnl = pd.read_excel(args.pnl) if args.pnl.endswith('.xlsx') else pd.read_csv(args.pnl)
        print(f"Loading GL: {args.gl}")
        df_gl = pd.read_excel(args.gl) if args.gl.endswith('.xlsx') else pd.read_csv(args.gl)
        # Detect periods from P&L columns (non-standard columns)
        standard_cols = {'Account Code', 'Account', 'Description', 'Account Name'}
        periods = [c for c in df_pnl.columns if c not in standard_cols]
    else:
        parser.error("Either --demo or both --pnl and --gl are required")
        return

    output_path, kpis, top_v, variances = build_report(
        df_pnl, df_gl, periods,
        entity_name=args.entity,
        currency=args.currency,
        industry=args.industry,
        output_path=args.output
    )

    print(f"\n{'='*60}")
    print(f"Report saved: {output_path}")
    print(f"{'='*60}")
    print("\nKey Findings:")
    for kpi in kpis:
        if kpi['pct_revenue'] is not None:
            print(f"  {kpi['metric']}: ${kpi['value']:,.0f} ({kpi['pct_revenue']:.1%}) {kpi['status']}")
        else:
            print(f"  {kpi['metric']}: ${kpi['value']:,.0f}")
    print(f"\nTop 5 Vendors:")
    for name, spend, pct in top_v[:5]:
        print(f"  {name}: ${spend:,.0f} ({pct:.1%})")
    flagged = [v for v in variances if v['flag']]
    if flagged:
        print(f"\nFlagged Variances ({len(flagged)}):")
        for v in flagged[:5]:
            print(f"  {v['commentary']}")


if __name__ == '__main__':
    main()
